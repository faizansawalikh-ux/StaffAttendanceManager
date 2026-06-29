"""
Excel file management for Staff Attendance Manager.
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import DATABASE_FILE, EXCEL_SHEETS
from utils import Logger, DateTimeUtil
from employee import Employee


class ExcelManager:
    """Manages Excel workbook operations."""

    def __init__(self, file_path: Path = DATABASE_FILE):
        """Initialize Excel manager."""
        self.file_path = file_path
        self.logger = Logger()

    def initialize_workbook(self) -> bool:
        """Create and initialize Excel workbook with all sheets."""
        try:
            wb = Workbook()
            wb.remove(wb.active)

            self._create_employees_sheet(wb)
            self._create_attendance_sheet(wb)
            self._create_movement_log_sheet(wb)
            self._create_daily_report_sheet(wb)
            self._create_monthly_report_sheet(wb)
            self._create_dashboard_data_sheet(wb)

            wb.save(self.file_path)
            self.logger.info(f"Workbook initialized: {self.file_path}", "ExcelManager")
            return True
        except Exception as e:
            self.logger.error("Failed to initialize workbook", "ExcelManager", exception=e)
            return False

    def _create_employees_sheet(self, wb: Workbook) -> None:
        """Create Employees sheet."""
        ws = wb.create_sheet(EXCEL_SHEETS["employees"])
        headers = [
            "Employee ID", "Name", "Department", "Designation",
            "Phone", "Email", "Joining Date", "Status", "Photo Path",
            "Notes", "Created At", "Updated At"
        ]
        self._write_headers(ws, headers)

    def _create_attendance_sheet(self, wb: Workbook) -> None:
        """Create Attendance sheet."""
        ws = wb.create_sheet(EXCEL_SHEETS["attendance"])
        headers = [
            "Employee ID", "Name", "Date", "Time In", "Time Out",
            "Status", "Duration (Hours)", "Present", "Created At"
        ]
        self._write_headers(ws, headers)

    def _create_movement_log_sheet(self, wb: Workbook) -> None:
        """Create MovementLog sheet."""
        ws = wb.create_sheet(EXCEL_SHEETS["movement_log"])
        headers = [
            "Log ID", "Employee ID", "Name", "Date", "Time",
            "Action", "Status", "Notes", "Created At"
        ]
        self._write_headers(ws, headers)

    def _create_daily_report_sheet(self, wb: Workbook) -> None:
        """Create DailyReport sheet."""
        ws = wb.create_sheet(EXCEL_SHEETS["daily_report"])
        headers = [
            "Date", "Total Employees", "Present", "Absent",
            "Inside Office", "Outside Office", "On Leave", "Generated At"
        ]
        self._write_headers(ws, headers)

    def _create_monthly_report_sheet(self, wb: Workbook) -> None:
        """Create MonthlyReport sheet."""
        ws = wb.create_sheet(EXCEL_SHEETS["monthly_report"])
        headers = [
            "Employee ID", "Name", "Month", "Total Days", "Present Days",
            "Absent Days", "Attendance %", "Generated At"
        ]
        self._write_headers(ws, headers)

    def _create_dashboard_data_sheet(self, wb: Workbook) -> None:
        """Create DashboardData sheet."""
        ws = wb.create_sheet(EXCEL_SHEETS["dashboard_data"])
        headers = ["Metric", "Value", "Date", "Time"]
        self._write_headers(ws, headers)

    def _write_headers(self, ws, headers: List[str]) -> None:
        """Write header row with formatting."""
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 25

    def get_workbook(self) -> Optional[Workbook]:
        """Get workbook, create if not exists."""
        try:
            if not self.file_path.exists():
                self.initialize_workbook()

            return load_workbook(self.file_path)
        except Exception as e:
            self.logger.error("Failed to get workbook", "ExcelManager", exception=e)
            return None

    def save_workbook(self, wb: Workbook) -> bool:
        """Save workbook."""
        try:
            wb.save(self.file_path)
            self.logger.debug(f"Workbook saved: {self.file_path}", "ExcelManager")
            return True
        except Exception as e:
            self.logger.error("Failed to save workbook", "ExcelManager", exception=e)
            return False

    def add_employee_record(self, employee: Employee) -> bool:
        """Add employee record to sheet."""
        try:
            wb = self.get_workbook()
            if not wb:
                return False

            ws = wb[EXCEL_SHEETS["employees"]]
            row = ws.max_row + 1

            values = [
                employee.employee_id,
                employee.name,
                employee.department,
                employee.designation,
                employee.phone,
                employee.email,
                employee.joining_date,
                employee.status,
                employee.photo_path or "",
                employee.notes or "",
                employee.created_at,
                employee.updated_at
            ]

            for col_idx, value in enumerate(values, 1):
                ws.cell(row=row, column=col_idx).value = value

            self._auto_fit_columns(ws)
            return self.save_workbook(wb)
        except Exception as e:
            self.logger.error("Failed to add employee record", "ExcelManager", exception=e)
            return False

    def get_employees(self) -> List[Dict[str, Any]]:
        """Get all employees from sheet."""
        try:
            wb = self.get_workbook()
            if not wb:
                return []

            ws = wb[EXCEL_SHEETS["employees"]]
            employees = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue

                emp_dict = {
                    "employee_id": row[0],
                    "name": row[1],
                    "department": row[2],
                    "designation": row[3],
                    "phone": row[4],
                    "email": row[5],
                    "joining_date": row[6],
                    "status": row[7],
                    "photo_path": row[8],
                    "notes": row[9],
                    "created_at": row[10],
                    "updated_at": row[11],
                }
                employees.append(emp_dict)

            return employees
        except Exception as e:
            self.logger.error("Failed to get employees", "ExcelManager", exception=e)
            return []

    def update_employee_record(self, employee_id: str, employee: Employee) -> bool:
        """Update employee record."""
        try:
            wb = self.get_workbook()
            if not wb:
                return False

            ws = wb[EXCEL_SHEETS["employees"]]

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                if row[0].value == employee_id:
                    values = [
                        employee.employee_id,
                        employee.name,
                        employee.department,
                        employee.designation,
                        employee.phone,
                        employee.email,
                        employee.joining_date,
                        employee.status,
                        employee.photo_path or "",
                        employee.notes or "",
                        employee.created_at,
                        DateTimeUtil.get_current_datetime()
                    ]

                    for col_idx, value in enumerate(values, 1):
                        row[col_idx - 1].value = value

                    self._auto_fit_columns(ws)
                    return self.save_workbook(wb)

            return False
        except Exception as e:
            self.logger.error("Failed to update employee record", "ExcelManager", exception=e)
            return False

    def delete_employee_record(self, employee_id: str) -> bool:
        """Delete employee record."""
        try:
            wb = self.get_workbook()
            if not wb:
                return False

            ws = wb[EXCEL_SHEETS["employees"]]

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                if row[0].value == employee_id:
                    ws.delete_rows(row_idx)
                    return self.save_workbook(wb)

            return False
        except Exception as e:
            self.logger.error("Failed to delete employee record", "ExcelManager", exception=e)
            return False

    def add_attendance_record(self, record: Dict[str, Any]) -> bool:
        """Add attendance record."""
        try:
            wb = self.get_workbook()
            if not wb:
                return False

            ws = wb[EXCEL_SHEETS["attendance"]]
            row = ws.max_row + 1

            values = [
                record.get("employee_id"),
                record.get("name"),
                record.get("date"),
                record.get("time_in"),
                record.get("time_out", ""),
                record.get("status"),
                record.get("duration", ""),
                record.get("present", True),
                DateTimeUtil.get_current_datetime()
            ]

            for col_idx, value in enumerate(values, 1):
                ws.cell(row=row, column=col_idx).value = value

            self._auto_fit_columns(ws)
            return self.save_workbook(wb)
        except Exception as e:
            self.logger.error("Failed to add attendance record", "ExcelManager", exception=e)
            return False

    def get_attendance_records(self, employee_id: Optional[str] = None, date: Optional[str] = None) -> List[Dict]:
        """Get attendance records with optional filtering."""
        try:
            wb = self.get_workbook()
            if not wb:
                return []

            ws = wb[EXCEL_SHEETS["attendance"]]
            records = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue

                if employee_id and row[0] != employee_id:
                    continue

                if date and row[2] != date:
                    continue

                record_dict = {
                    "employee_id": row[0],
                    "name": row[1],
                    "date": row[2],
                    "time_in": row[3],
                    "time_out": row[4],
                    "status": row[5],
                    "duration": row[6],
                    "present": row[7],
                    "created_at": row[8],
                }
                records.append(record_dict)

            return records
        except Exception as e:
            self.logger.error("Failed to get attendance records", "ExcelManager", exception=e)
            return []

    def add_movement_log(self, log: Dict[str, Any]) -> bool:
        """Add movement log record."""
        try:
            wb = self.get_workbook()
            if not wb:
                return False

            ws = wb[EXCEL_SHEETS["movement_log"]]
            row = ws.max_row + 1

            log_id = f"LOG_{row}_{DateTimeUtil.get_current_time().replace(':', '')}"
            values = [
                log_id,
                log.get("employee_id"),
                log.get("name"),
                log.get("date"),
                log.get("time"),
                log.get("action"),
                log.get("status"),
                log.get("notes", ""),
                DateTimeUtil.get_current_datetime()
            ]

            for col_idx, value in enumerate(values, 1):
                ws.cell(row=row, column=col_idx).value = value

            self._auto_fit_columns(ws)
            return self.save_workbook(wb)
        except Exception as e:
            self.logger.error("Failed to add movement log", "ExcelManager", exception=e)
            return False

    def get_movement_logs(self, employee_id: Optional[str] = None, date: Optional[str] = None) -> List[Dict]:
        """Get movement logs with optional filtering."""
        try:
            wb = self.get_workbook()
            if not wb:
                return []

            ws = wb[EXCEL_SHEETS["movement_log"]]
            logs = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue

                if employee_id and row[1] != employee_id:
                    continue

                if date and row[3] != date:
                    continue

                log_dict = {
                    "log_id": row[0],
                    "employee_id": row[1],
                    "name": row[2],
                    "date": row[3],
                    "time": row[4],
                    "action": row[5],
                    "status": row[6],
                    "notes": row[7],
                    "created_at": row[8],
                }
                logs.append(log_dict)

            return logs
        except Exception as e:
            self.logger.error("Failed to get movement logs", "ExcelManager", exception=e)
            return []

    def _auto_fit_columns(self, ws) -> None:
        """Auto-fit column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
